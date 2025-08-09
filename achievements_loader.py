from openpyxl import load_workbook

async def load_achievements_from_excel(file_path="achievements.xlsx"):
    wb = load_workbook(file_path)
    ws = wb.active
    achievements = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        category, code, title, description = row
        if not code:
            continue
        achievements.append({
            "category": category,
            "code": code,
            "title": title,
            "description": description
        })

    return achievements
